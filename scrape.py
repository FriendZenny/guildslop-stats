#!/usr/bin/env python3
"""
Guild stats scraper — rockymoon.com → guild_stats.xlsx

Run whenever you want a new snapshot:
    python3 scrape.py

Each run appends to the Data sheet and rewrites the Changes sheet
so you can see gains/losses since the last run.
"""

import re
import sys
import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

GUILD_URL = "https://www.rockymoon.com/bob/Charts/Borealis/Guild/Index/91"
OUTPUT_FILE = Path("guild_stats.xlsx")
HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; guild-stats/1.0)"}

# Colours (ARGB)
COL_HEADER   = "FF4472C4"
COL_ALT_ROW  = "FFD9E1F2"
COL_GREEN    = "FFC6EFCE"
COL_RED      = "FFFFC7CE"
COL_NEW      = "FFFFFF00"


# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

def _parse_num(text: str):
    """'1,845' → 1845  |  '28.85' → 28.85  |  anything else → None"""
    t = text.strip().replace(",", "")
    if not t or t == "-":
        return None
    try:
        return float(t) if "." in t else int(t)
    except ValueError:
        return None


def scrape() -> list[dict]:
    resp = requests.get(GUILD_URL, headers=HTTP_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Find the table whose thead includes "Player" and "Economy"
    # Note: on this site <th> are direct children of <thead>, not inside a <tr>
    member_table = None
    for tbl in soup.find_all("table"):
        thead = tbl.find("thead")
        if not thead:
            continue
        ths_text = [th.get_text(strip=True).lower() for th in thead.find_all("th")]
        if "player" in ths_text and "economy" in ths_text:
            member_table = tbl
            break

    if not member_table:
        sys.exit("ERROR: Could not find the member stats table. The page layout may have changed.")

    # Map header name → column index
    thead = member_table.find("thead")
    ths = [th.get_text(strip=True).lower() for th in thead.find_all("th")]
    col = {name: idx for idx, name in enumerate(ths)}

    player_link_re = re.compile(r"/Player/Index/(\d+)$")

    members = []
    tbody = member_table.find("tbody") or member_table
    for row in tbody.find_all("tr"):
        tds = row.find_all("td")
        if len(tds) < 4:
            continue

        # Extract player name + ID from the first link matching the player URL pattern
        player_id, player_name = None, None
        for td in tds:
            link = td.find("a", href=player_link_re)
            if link:
                m = player_link_re.search(link["href"])
                player_id = int(m.group(1)) if m else None
                player_name = link.get_text(strip=True)
                break

        if not player_name:
            continue

        def cell_num(key):
            idx = col.get(key)
            if idx is None or idx >= len(tds):
                return None
            return _parse_num(tds[idx].get_text(strip=True))

        members.append({
            "player_id": player_id,
            "player":     player_name,
            "level":      cell_num("level"),
            "economy":    cell_num("economy"),
            "fleet":      cell_num("fleet"),
            "technology": cell_num("technology"),
            "experience": cell_num("experience"),
        })

    return members


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

DATA_HEADERS = ["Run", "Player ID", "Player", "Level", "Economy", "Fleet", "Technology", "Experience"]


def _header_style(cell, color=COL_HEADER):
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _get_or_create_sheet(wb, name, index=None):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name, index)
    return ws


def get_last_snapshot(wb) -> tuple[datetime.datetime | None, dict]:
    """
    Return (last_run_datetime, {player_id: stat_dict}) for the most recent run
    recorded in the Data sheet.  Returns (None, {}) if no prior data exists.
    """
    if "Data" not in wb.sheetnames:
        return None, {}

    ws = wb["Data"]
    if ws.max_row < 2:
        return None, {}

    last_dt = None
    snapshot: dict[int, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        run_dt, pid, player, level, economy, fleet, technology, experience = row[:8]
        if run_dt is None:
            continue
        # openpyxl returns datetime for datetime cells
        if isinstance(run_dt, datetime.date) and not isinstance(run_dt, datetime.datetime):
            run_dt = datetime.datetime(run_dt.year, run_dt.month, run_dt.day)

        if last_dt is None or run_dt > last_dt:
            last_dt = run_dt
            snapshot = {}

        if run_dt == last_dt and pid is not None:
            snapshot[int(pid)] = {
                "player":     player,
                "level":      level,
                "economy":    economy,
                "fleet":      fleet,
                "technology": technology,
                "experience": experience,
            }

    return last_dt, snapshot


def append_data(wb, run_dt: datetime.datetime, members: list[dict]):
    ws = _get_or_create_sheet(wb, "Data")

    # Write headers on first use
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for c, h in enumerate(DATA_HEADERS, 1):
            _header_style(ws.cell(1, c, h))
        ws.freeze_panes = "A2"

    for m in members:
        ws.append([
            run_dt,
            m["player_id"],
            m["player"],
            m["level"],
            m["economy"],
            m["fleet"],
            m["technology"],
            m["experience"],
        ])

    # Format the Run column as a readable datetime
    dt_col = 1
    fmt = "yyyy-mm-dd hh:mm"
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, dt_col).number_format = fmt


def _delta_cell(ws, row, col, delta):
    """Write a numeric delta with green/red background."""
    c = ws.cell(row, col, delta)
    c.alignment = Alignment(horizontal="center")
    if delta is None:
        c.value = ""
        return
    if delta > 0:
        c.fill = PatternFill("solid", fgColor=COL_GREEN)
    elif delta < 0:
        c.fill = PatternFill("solid", fgColor=COL_RED)


def _safe_delta(cur, prev):
    if cur is None or prev is None:
        return None
    return cur - prev


def write_changes(wb, run_dt: datetime.datetime, last_dt: datetime.datetime | None,
                  current: list[dict], last_snap: dict):
    # Always rebuild this sheet from scratch
    if "Changes" in wb.sheetnames:
        del wb["Changes"]
    ws = wb.create_sheet("Changes", 0)   # first tab

    # ── Title ──────────────────────────────────────────────────────────────
    if last_dt:
        days = (run_dt.date() - last_dt.date()).days
        title = (f"Changes  ·  {last_dt.strftime('%Y-%m-%d')} → "
                 f"{run_dt.strftime('%Y-%m-%d')}  ({days} day{'s' if days != 1 else ''})")
    else:
        title = f"First snapshot  ·  {run_dt.strftime('%Y-%m-%d %H:%M')}"

    ws.cell(1, 1, title).font = Font(bold=True, size=13)
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 22

    # ── Column headers ─────────────────────────────────────────────────────
    COLS = [
        "Player", "Level", "ΔLevel",
        "Economy", "ΔEconomy",
        "Fleet", "ΔFleet",
        "Technology", "ΔTech",
        "Experience", "ΔExp",
        "Status",
    ]
    for c, h in enumerate(COLS, 1):
        _header_style(ws.cell(2, c, h))
    ws.freeze_panes = "A3"

    # ── Data rows ──────────────────────────────────────────────────────────
    # Sort current members by economy descending (highest first)
    sorted_current = sorted(current, key=lambda x: x["economy"] or 0, reverse=True)

    for r_idx, m in enumerate(sorted_current, start=3):
        pid  = m["player_id"]
        prev = last_snap.get(pid)

        # Alternate row shading on non-delta columns
        alt = (r_idx % 2 == 0)
        def plain_cell(row, col, value):
            c = ws.cell(row, col, value)
            if alt:
                c.fill = PatternFill("solid", fgColor=COL_ALT_ROW)
            return c

        plain_cell(r_idx, 1, m["player"])
        plain_cell(r_idx, 2, m["level"])
        _delta_cell(ws, r_idx, 3,  _safe_delta(m["level"],      prev["level"]      if prev else None))
        plain_cell(r_idx, 4, m["economy"])
        _delta_cell(ws, r_idx, 5,  _safe_delta(m["economy"],    prev["economy"]    if prev else None))
        plain_cell(r_idx, 6, m["fleet"])
        _delta_cell(ws, r_idx, 7,  _safe_delta(m["fleet"],      prev["fleet"]      if prev else None))
        plain_cell(r_idx, 8, m["technology"])
        _delta_cell(ws, r_idx, 9,  _safe_delta(m["technology"], prev["technology"] if prev else None))
        plain_cell(r_idx, 10, m["experience"])
        _delta_cell(ws, r_idx, 11, _safe_delta(m["experience"], prev["experience"] if prev else None))

        status_cell = ws.cell(r_idx, 12)
        if prev is None:
            status_cell.value = "NEW"
            status_cell.font = Font(bold=True, color="FF00B050")
            status_cell.fill = PatternFill("solid", fgColor=COL_NEW)
        else:
            plain_cell(r_idx, 12, "")

    # ── Departed members ───────────────────────────────────────────────────
    current_ids = {m["player_id"] for m in current}
    departed = [(pid, snap) for pid, snap in last_snap.items() if pid not in current_ids]
    if departed:
        r_idx = 3 + len(sorted_current)
        for pid, prev in sorted(departed, key=lambda x: x[1]["player"]):
            for c in range(1, 13):
                ws.cell(r_idx, c).font = Font(italic=True, color="FF808080")
            ws.cell(r_idx, 1, prev["player"])
            ws.cell(r_idx, 12, "LEFT")
            r_idx += 1

    # ── Column widths ──────────────────────────────────────────────────────
    widths = [20, 8, 8, 10, 10, 10, 10, 12, 10, 12, 10, 8]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ---------------------------------------------------------------------------
# Terminal summary
# ---------------------------------------------------------------------------

def print_summary(run_dt, last_dt, current, last_snap):
    if not last_dt:
        print("First snapshot — no previous data to compare.")
        return

    days = (run_dt.date() - last_dt.date()).days
    print(f"\n{'─'*80}")
    print(f"  Changes: {last_dt.strftime('%Y-%m-%d')} → {run_dt.strftime('%Y-%m-%d')}  ({days} days)")
    print(f"{'─'*80}")
    print(f"  {'Player':<20} {'Level':>7}  {'Economy':>10}  {'Fleet':>11}  {'Tech':>11}  {'XP':>11}")
    print(f"  {'─'*18}  {'─'*7}  {'─'*10}  {'─'*11}  {'─'*11}  {'─'*11}")

    def fmt(val, prev_val, float_fmt=False):
        if val is None:
            return "     N/A"
        if prev_val is None:
            return "     NEW"
        d = val - prev_val
        if float_fmt:
            return f"{d:+.2f}"
        return f"{d:+,}"

    sorted_current = sorted(current, key=lambda x: x["economy"] or 0, reverse=True)
    for m in sorted_current:
        prev = last_snap.get(m["player_id"])
        pv = lambda k: prev[k] if prev else None
        print(
            f"  {m['player']:<20} "
            f"{fmt(m['level'],      pv('level'),      True):>7}  "
            f"{fmt(m['economy'],    pv('economy')):>10}  "
            f"{fmt(m['fleet'],      pv('fleet')):>11}  "
            f"{fmt(m['technology'], pv('technology')):>11}  "
            f"{fmt(m['experience'], pv('experience')):>11}"
        )

    current_ids = {m["player_id"] for m in current}
    for pid, prev in last_snap.items():
        if pid not in current_ids:
            print(f"  {prev['player']:<20}  (left guild)")

    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    run_dt = datetime.datetime.now().replace(second=0, microsecond=0)

    print(f"Fetching {GUILD_URL} ...")
    members = scrape()
    print(f"  → {len(members)} members found.")

    if OUTPUT_FILE.exists():
        wb = openpyxl.load_workbook(OUTPUT_FILE)
    else:
        wb = openpyxl.Workbook()
        # Remove the blank default sheet
        del wb[wb.sheetnames[0]]

    last_dt, last_snap = get_last_snapshot(wb)

    append_data(wb, run_dt, members)
    write_changes(wb, run_dt, last_dt, members, last_snap)

    wb.save(OUTPUT_FILE)

    print_summary(run_dt, last_dt, members, last_snap)
    print(f"Saved → {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
